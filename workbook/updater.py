from __future__ import annotations

import json
import os
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from results.models import ExecutionResult, WORKBOOK_EXECUTION_FIELDS
from workbook.models import EXECUTION_LOG_SHEET
from workbook.parser import load_case_index


EXECUTION_ID_RE = re.compile(r"^EXE-(\d+)$")
EXECUTION_ID_COLUMN = 1
CASE_ID_COLUMN = 2


def _value(value: object) -> str:
    return "" if value is None else str(value).strip()


def next_execution_number(ws, *, min_row: int) -> int:
    max_number = 0
    for row in ws.iter_rows(min_row=min_row, min_col=1, max_col=1, values_only=True):
        cell = _value(row[0])
        match = EXECUTION_ID_RE.match(cell)
        if match:
            max_number = max(max_number, int(match.group(1)))
    return max_number + 1


def first_available_row(ws, *, min_row: int) -> int:
    row_number = min_row
    while _value(ws.cell(row=row_number, column=CASE_ID_COLUMN).value) != "":
        row_number += 1
    return row_number


def _cell_has_formula(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def find_execution_log_header_row(ws) -> int:
    for row_number in range(1, min(ws.max_row, 20) + 1):
        headers = [
            _value(ws.cell(row=row_number, column=idx).value)
            for idx in range(1, len(WORKBOOK_EXECUTION_FIELDS) + 1)
        ]
        if headers == WORKBOOK_EXECUTION_FIELDS:
            return row_number
    raise ValueError(f"{EXECUTION_LOG_SHEET} header row was not found in the first 20 rows.")


def append_results_to_workbook(
    workbook_path: str | Path,
    results: list[ExecutionResult],
    *,
    output_path: str | Path | None = None,
    in_place: bool = False,
    ledger_path: str | Path | None = None,
) -> Path:
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    if not results:
        raise ValueError("No execution results were provided.")

    run_ids = {result.run_id for result in results}
    if len(run_ids) != 1:
        raise ValueError(f"Expected exactly one run_id per import, found: {sorted(run_ids)}")
    run_id = next(iter(run_ids))

    ledger_path = Path(ledger_path or workbook_path.with_suffix(".import-ledger.json"))
    imported_runs = _read_ledger(ledger_path)
    if run_id in imported_runs:
        raise ValueError(f"Run {run_id} was already imported according to {ledger_path}.")

    case_index = load_case_index(workbook_path)
    for result in results:
        workbook_case = case_index.get(result.case_id)
        if workbook_case is None:
            raise ValueError(f"Case ID {result.case_id} does not exist in the workbook.")
        if workbook_case.module != result.module:
            raise ValueError(
                f"Module mismatch for {result.case_id}: result={result.module!r}, workbook={workbook_case.module!r}."
            )

    if in_place:
        backup_path = workbook_path.with_suffix(f".backup-{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        shutil.copy2(workbook_path, backup_path)
        destination = workbook_path
    else:
        destination = Path(output_path) if output_path else workbook_path.with_name(
            f"{workbook_path.stem}_with_automation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

    temp_destination = destination.with_name(f"{destination.stem}.tmp{destination.suffix}")
    written_rows: list[tuple[int, str]] = []
    wb = load_workbook(workbook_path)
    try:
        if EXECUTION_LOG_SHEET not in wb.sheetnames:
            raise ValueError(f"Workbook is missing required sheet: {EXECUTION_LOG_SHEET}")
        ws = wb[EXECUTION_LOG_SHEET]
        header_row = find_execution_log_header_row(ws)
        execution_number = next_execution_number(ws, min_row=header_row + 1)
        row_number = first_available_row(ws, min_row=header_row + 1)

        for result in results:
            row = result.to_workbook_row(f"EXE-{execution_number:04d}")
            for column_number, field_name in enumerate(WORKBOOK_EXECUTION_FIELDS, start=1):
                cell = ws.cell(row=row_number, column=column_number)
                if column_number == EXECUTION_ID_COLUMN and _cell_has_formula(cell):
                    continue
                cell.value = row[field_name]
            written_rows.append((row_number, result.case_id))
            execution_number += 1
            row_number = first_available_row(ws, min_row=row_number + 1)

        destination.parent.mkdir(parents=True, exist_ok=True)
        wb.save(temp_destination)
    finally:
        wb.close()

    validation_wb = load_workbook(temp_destination, read_only=True, data_only=False)
    try:
        if EXECUTION_LOG_SHEET not in validation_wb.sheetnames:
            raise ValueError(f"Saved workbook is missing required sheet: {EXECUTION_LOG_SHEET}")
        validation_ws = validation_wb[EXECUTION_LOG_SHEET]
        find_execution_log_header_row(validation_ws)
        for row_number, case_id in written_rows:
            saved_case_id = _value(validation_ws.cell(row=row_number, column=CASE_ID_COLUMN).value)
            if saved_case_id != case_id:
                raise ValueError(
                    f"Saved workbook validation failed: expected Case ID {case_id} at row {row_number}, "
                    f"found {saved_case_id or '<blank>'}."
                )
    finally:
        validation_wb.close()

    os.replace(temp_destination, destination)

    imported_runs.add(run_id)
    _write_ledger(ledger_path, imported_runs)
    return destination


def _read_ledger(path: Path) -> set[str]:
    if not path.exists():
        return set()
    payload = json.loads(path.read_text(encoding="utf-8"))
    return set(payload.get("imported_run_ids", []))


def _write_ledger(path: Path, imported_runs: set[str]) -> None:
    path.write_text(
        json.dumps({"imported_run_ids": sorted(imported_runs)}, indent=2),
        encoding="utf-8",
    )
