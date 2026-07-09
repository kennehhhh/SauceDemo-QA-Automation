from __future__ import annotations

from pathlib import Path
from collections import defaultdict
from typing import Iterable

from openpyxl import load_workbook

from workbook.module_registry import MODULES_BY_CODE, WORKBOOK_MODULES
from workbook.models import MODULE_SHEET_SUFFIX, WorkbookCase


EXPECTED_MODULE_HEADERS = [
    "Case ID",
    "Test Category",
    "Module",
    "Test Case / Scenario",
    "Preconditions",
    "Test Data",
    "Test Steps",
    "Expected Result",
    "Applicable User",
    "Browser Coverage",
    "Remarks",
    "Automation Candidate",
]


def _value(value: object) -> str:
    return "" if value is None else str(value).strip()


def _row_values(row: tuple[object, ...]) -> list[str]:
    values = [_value(cell) for cell in row[: len(EXPECTED_MODULE_HEADERS)]]
    return values + [""] * (len(EXPECTED_MODULE_HEADERS) - len(values))


def module_sheet_names(workbook_path: str | Path) -> list[str]:
    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        return [name for name in wb.sheetnames if name.endswith(MODULE_SHEET_SUFFIX)]
    finally:
        wb.close()


def iter_cases(workbook_path: str | Path) -> Iterable[WorkbookCase]:
    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        for sheet_name in wb.sheetnames:
            if not sheet_name.endswith(MODULE_SHEET_SUFFIX):
                continue
            ws = wb[sheet_name]
            header_row_number = None
            for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                values = _row_values(row)
                if values == EXPECTED_MODULE_HEADERS:
                    header_row_number = row_number
                    break
            if header_row_number is None:
                continue

            for row in ws.iter_rows(min_row=header_row_number + 1, values_only=True):
                values = _row_values(row)
                if not values[0] or values[0] == "Case ID":
                    continue
                if "-" not in values[0]:
                    continue
                yield WorkbookCase(
                    case_id=values[0],
                    category=values[1],
                    module=values[2],
                    scenario=values[3],
                    preconditions=values[4],
                    test_data=values[5],
                    steps=values[6],
                    expected_result=values[7],
                    applicable_user=values[8],
                    browser_coverage=values[9],
                    remarks=values[10],
                    automation_candidate=values[11],
                    sheet_name=sheet_name,
                )
    finally:
        wb.close()


def load_case_index(workbook_path: str | Path) -> dict[str, WorkbookCase]:
    cases: dict[str, WorkbookCase] = {}
    duplicates: list[str] = []
    for case in iter_cases(workbook_path):
        if case.case_id in cases:
            duplicates.append(case.case_id)
        cases[case.case_id] = case
    if duplicates:
        raise ValueError(f"Duplicate workbook Case IDs found: {sorted(set(duplicates))}")
    return cases


def load_cases_by_module(workbook_path: str | Path) -> dict[str, list[WorkbookCase]]:
    grouped: dict[str, list[WorkbookCase]] = defaultdict(list)
    for case in load_case_index(workbook_path).values():
        grouped[case.module].append(case)
    return dict(grouped)


def validate_workbook_contract(workbook_path: str | Path, *, expected_total: int | None = 239) -> list[str]:
    errors: list[str] = []
    sheet_names = set(module_sheet_names(workbook_path))
    for module in WORKBOOK_MODULES:
        if module.sheet_name not in sheet_names:
            errors.append(f"Missing module sheet: {module.sheet_name}")

    cases = load_case_index(workbook_path)
    if expected_total is not None and len(cases) != expected_total:
        errors.append(f"Expected {expected_total} workbook cases, found {len(cases)}")

    for case in cases.values():
        prefix = case.case_id.split("-", 1)[0]
        module = MODULES_BY_CODE.get(prefix)
        if module is None:
            errors.append(f"{case.case_id}: unknown module prefix {prefix}")
            continue
        if case.sheet_name != module.sheet_name:
            errors.append(f"{case.case_id}: expected sheet {module.sheet_name}, found {case.sheet_name}")
    return errors
