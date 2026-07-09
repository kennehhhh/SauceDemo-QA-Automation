from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from results.models import ExecutionResult, WORKBOOK_EXECUTION_FIELDS
from workbook.updater import append_results_to_workbook, find_execution_log_header_row, first_available_row, next_execution_number


def make_result(case_id: str = "LG-0001", module: str = "Login & Session", run_id: str = "RUN-20260709T103000-12345678"):
    return ExecutionResult(
        case_id=case_id,
        module=module,
        test_user="N/A",
        browser="Chrome",
        actual_result="Username input was visible.",
        status="Passed",
        related_defect_id="",
        tester="Tester",
        execution_date="2026-07-09 13:00:00 +0800",
        remarks=f"Automated by Selenium + pytest; run_id={run_id}",
        run_id=run_id,
    )


def make_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Login & Session - TC"
    ws.append(["Login & Session Test Cases"])
    ws.append(["Module Code: LG"])
    ws.append(
        [
            "Case ID",
            "Test Category",
            "Module",
            "Test Case / Scenario",
            "Preconditions",
            "Test Data",
            "Test Steps",
            "Expected Result",
            "Applicable User",
            "Browser Coverage",
            "Remarks",
            "Automation Candidate",
        ]
    )
    ws.append(["LG-0001", "Unit", "Login & Session", "Username field is visible", "", "", "", "", "N/A", "Primary Browser", "", "Yes"])

    log = wb.create_sheet("Test Execution Log")
    log.append(["Test Execution Log"])
    log.append(["Record only actual executions."])
    log.append([])
    log.append(WORKBOOK_EXECUTION_FIELDS)
    wb.save(path)
    wb.close()
    return path


def make_formula_prefilled_workbook(path: Path) -> Path:
    make_workbook(path)
    wb = load_workbook(path)
    try:
        ws = wb["Test Execution Log"]
        for row_number in range(5, 1006):
            ws.cell(row=row_number, column=1).value = f'=IF(B{row_number}="","","EXE-"&TEXT(ROW()-4,"0000"))'
        wb.save(path)
    finally:
        wb.close()
    return path


def test_header_row_4_detection_and_next_execution(tmp_path):
    path = make_workbook(tmp_path / "source.xlsx")
    wb = load_workbook(path)
    try:
        ws = wb["Test Execution Log"]
        assert find_execution_log_header_row(ws) == 4
        assert first_available_row(ws, min_row=5) == 5
        assert next_execution_number(ws, min_row=5) == 1
    finally:
        wb.close()


def test_append_continues_execution_id_and_blocks_duplicate_run(tmp_path):
    source = make_workbook(tmp_path / "source.xlsx")
    output = tmp_path / "out.xlsx"
    ledger = tmp_path / "ledger.json"

    append_results_to_workbook(source, [make_result()], output_path=output, ledger_path=ledger)

    wb = load_workbook(output, read_only=True)
    try:
        ws = wb["Test Execution Log"]
        assert ws.cell(row=5, column=1).value == "EXE-0001"
        assert ws.cell(row=5, column=2).value == "LG-0001"
    finally:
        wb.close()

    with pytest.raises(ValueError, match="already imported"):
        append_results_to_workbook(source, [make_result()], output_path=tmp_path / "out2.xlsx", ledger_path=ledger)


def test_append_uses_first_blank_case_id_row_and_preserves_execution_id_formula(tmp_path):
    source = make_formula_prefilled_workbook(tmp_path / "source.xlsx")
    output = tmp_path / "out.xlsx"

    wb = load_workbook(source, read_only=False, data_only=False)
    try:
        ws = wb["Test Execution Log"]
        assert ws.max_row == 1005
        assert first_available_row(ws, min_row=5) == 5
    finally:
        wb.close()

    append_results_to_workbook(source, [make_result()], output_path=output, ledger_path=tmp_path / "ledger.json")

    wb = load_workbook(output, read_only=False, data_only=False)
    try:
        ws = wb["Test Execution Log"]
        assert ws.cell(row=5, column=1).value == '=IF(B5="","","EXE-"&TEXT(ROW()-4,"0000"))'
        assert ws.cell(row=5, column=2).value == "LG-0001"
        assert ws.cell(row=1006, column=2).value is None
    finally:
        wb.close()


def test_invalid_case_id_aborts_append_without_output(tmp_path):
    source = make_workbook(tmp_path / "source.xlsx")
    output = tmp_path / "out.xlsx"

    with pytest.raises(ValueError, match="does not exist"):
        append_results_to_workbook(source, [make_result(case_id="LG-9999")], output_path=output, ledger_path=tmp_path / "ledger.json")

    assert not output.exists()


def test_module_mismatch_aborts_append_without_output(tmp_path):
    source = make_workbook(tmp_path / "source.xlsx")
    output = tmp_path / "out.xlsx"

    with pytest.raises(ValueError, match="Module mismatch"):
        append_results_to_workbook(source, [make_result(module="Wrong Module")], output_path=output, ledger_path=tmp_path / "ledger.json")

    assert not output.exists()
